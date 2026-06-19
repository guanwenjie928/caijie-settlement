"""
Excel 导出服务模块
==================
使用 openpyxl 生成标准 xlsx 文件，支持按筛选条件导出结算记录。
包含表头样式、金额格式化、状态中文映射等。
结算逻辑: 盈利(profit_rate) + 税费(tax_rate) + 结算给他人(剩余)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 状态中文映射
STATUS_MAP = {
    "paid": "已结清",
    "settling": "正在结算",
    "unpaid": "尚未结清",
}

# 表头定义 — 包含盈利、税费、结算金额四列金额
HEADERS = [
    ("序号", 8),
    ("人名", 12),
    ("公司名", 28),
    ("税号", 22),
    ("原始金额", 14),
    ("盈利比例", 10),
    ("盈利金额", 14),
    ("税费比例", 10),
    ("税费金额", 14),
    ("结算金额", 14),
    ("已结金额", 14),
    ("录入时间", 22),
    ("状态", 10),
    ("结清时间", 22),
    ("备注", 20),
    ("来源文件", 20),
]


def export_to_excel(records: list[dict]) -> bytes:
    """
    将结算记录列表导出为 Excel 文件。

    参数:
        records: 记录字典列表

    返回:
        xlsx 文件的二进制内容
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "结算记录"

    # ── 样式定义 ──────────────────────────────────
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 金额单元格背景色
    amount_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")      # 原始金额 — 浅黄
    profit_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")      # 盈利 — 浅蓝
    tax_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")         # 税费 — 浅橙
    settlement_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 结算 — 浅绿

    settled_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")    # 已结金额 — 浅蓝

    # ── 写表头 ──────────────────────────────────
    for col_idx, (title, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # ── 写数据行 ────────────────────────────────
    for row_idx, record in enumerate(records, 2):
        values = [
            row_idx - 1,  # 序号
            record.get("person_name", ""),
            record.get("company_name", ""),
            record.get("tax_number", ""),
            record.get("original_amount", 0),
            f'{record.get("profit_rate", 0.04) * 100:.1f}%',
            record.get("profit_amount", 0),
            f'{record.get("tax_rate", 0.01) * 100:.1f}%',
            record.get("tax_amount", 0),
            record.get("settlement_amount", 0),
            record.get("settled_amount", 0),
            record.get("entry_time", ""),
            STATUS_MAP.get(record.get("status", ""), ""),
            record.get("settled_time", "") or "",
            record.get("remark", ""),
            record.get("source_file", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

            # 金额列格式化与着色
            if col_idx == 5:    # 原始金额
                cell.number_format = '#,##0.00'
                cell.fill = amount_fill
            elif col_idx == 7:  # 盈利金额
                cell.number_format = '#,##0.00'
                cell.fill = profit_fill
            elif col_idx == 9:  # 税费金额
                cell.number_format = '#,##0.00'
                cell.fill = tax_fill
            elif col_idx == 10: # 结算金额
                cell.number_format = '#,##0.00'
                cell.fill = settlement_fill
            elif col_idx == 11: # 已结金额
                cell.number_format = '#,##0.00'
                cell.fill = settled_fill

    # ── 底部统计行 ──────────────────────────────
    summary_row = len(records) + 2
    total_original = sum(r.get("original_amount", 0) or 0 for r in records)
    total_profit = sum(r.get("profit_amount", 0) or 0 for r in records)
    total_tax = sum(r.get("tax_amount", 0) or 0 for r in records)
    total_settlement = sum(r.get("settlement_amount", 0) or 0 for r in records)
    total_settled = sum(r.get("settled_amount", 0) or 0 for r in records)
    unpaid_count = sum(1 for r in records if r.get("status") == "unpaid")
    settling_count = sum(1 for r in records if r.get("status") == "settling")
    paid_count = sum(1 for r in records if r.get("status") == "paid")

    # 合计标签（合并前4列）
    cell = ws.cell(row=summary_row, column=1, value="合计")
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)

    # 合计原始金额
    cell = ws.cell(row=summary_row, column=5, value=total_original)
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = amount_fill

    # 比例列留空
    for col in (6, 8):
        cell = ws.cell(row=summary_row, column=col, value="")
        cell.border = thin_border

    # 合计盈利金额
    cell = ws.cell(row=summary_row, column=7, value=total_profit)
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = profit_fill

    # 合计税费金额
    cell = ws.cell(row=summary_row, column=9, value=total_tax)
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = tax_fill

    # 合计结算金额
    cell = ws.cell(row=summary_row, column=10, value=total_settlement)
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = settlement_fill

    # 合计已结金额
    cell = ws.cell(row=summary_row, column=11, value=total_settled)
    cell.font = Font(name="微软雅黑", size=11, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    cell.number_format = '#,##0.00'
    cell.fill = settled_fill

    # 统计信息
    info_text = f"已结清 {paid_count} 笔 / 正在结算 {settling_count} 笔 / 未结清 {unpaid_count} 笔 / 共 {len(records)} 笔"
    cell = ws.cell(row=summary_row, column=12, value=info_text)
    cell.font = Font(name="微软雅黑", size=10, color="666666")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=summary_row, start_column=12, end_row=summary_row, end_column=16)

    # ── 冻结首行 ────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 输出到内存 ──────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
